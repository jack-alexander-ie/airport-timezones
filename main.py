#!/usr/bin/env python3
# -*- coding: latin-1 -*-

import os
import sys
import datetime
import pytz
import json

import openpyxl
from openpyxl import load_workbook
from timezonefinder import TimezoneFinder

# This has to be kept updated manually until a better way is found...
updated_timezones = {
    "Asia/Qostanay": "Asia/Qyzylorda"
}


def get_timezone_info(lat, lon):

    # Allows timezone info to be found based on longitude/latitude coordinates
    timezone = TimezoneFinder().certain_timezone_at(lng=lon, lat=lat)

    # Check to see if timezone has been recently updated
    if timezone in updated_timezones:
        timezone = updated_timezones[timezone]

    return timezone


def get_offset(timezone):

    # Check if timezone is listed - not always the case
    if timezone == 'NA' or timezone is None:
        offset, offset_secs = 'NA', ''
    else:

        offset = pytz.timezone(timezone).localize(datetime.datetime(2011, 1, 1)).strftime('%z')

        # Format offset to conform with our DB
        offset = offset[:3] + ":" + offset[3:5] + ":00"
        offset_secs = int(pytz.timezone(timezone)._transition_info[-2][1].total_seconds())

        # Check if offset seconds is 0 or not
        if offset_secs == 0:
            offset_secs = ''
        else:
            offset_secs = str(offset_secs)

    return offset, offset_secs


def get_dst_info(timezone, dst_year) -> list:

    dst_dates = []

    # Grabs all DST dates for each entry
    dst_array = pytz.timezone(timezone)._utc_transition_times

    # Searches through all DST dates for each entry and grabs the relevant start and end dates
    for entry in dst_array:

        year = entry.strftime('%Y')

        if year == dst_year:
            date = entry.strftime('%d/%m/%y')
            dst_dates.append(date)

    return dst_dates


def get_airport_data(file_name, dst_year) -> dict:

    """
    Reads all airport data in from a structured spreadsheet.

    :param file_name: the string file name to the Excel spreadsheet
    :param dst_year: the year to collect the DST info for
    :return: a dictionary with all the info for creating other sheets
    """

    print('Collecting airport information...')

    # Get the Excel file with the airport info, load the workbook and its sheet
    xlsx_file = os.getcwd() + '/' + file_name
    wb = load_workbook(filename=xlsx_file, data_only=True)
    master_sheet = wb['MASTER_LIST']

    # Will hold list of airports and IATA codes
    airports = {}

    # Collect info on each airport in the DB
    for row in master_sheet.iter_rows(row_offset=1):

        # Grab relevant info from the master list
        iata, airport_name, latitude, longitude = row[0].value, row[1].value, row[3].value, row[4].value

        if iata:

            # Allows timezone info to be found based on longitude/latitude coordinates
            timezone = get_timezone_info(latitude, longitude)

            # Gets the time offset in seconds
            offset, offset_secs = get_offset(timezone)

            # Use to collect DST start and DST end
            dst_dates = get_dst_info(timezone, dst_year)

            # Add it to the dict
            airports[iata] = [airport_name, timezone, str(latitude), str(longitude), offset, offset_secs, dst_dates]

    wb.save(xlsx_file)

    return airports


def open_json_db() -> dict:
    with open('airport-db.json') as json_file:
        json_data = json.load(json_file)
        data = json.loads(json_data)
        return data


def update_json_db(data):
    json_data = json.dumps(data, indent=4)
    with open('airport-db.json', 'w') as outfile:
        json.dump(json_data, outfile)


def get_from_db(iata):
    json_db_data = open_json_db()
    if iata in json_db_data:
        print('Data:', json_db_data[iata])
        return json_db_data[iata]
    else:
        print('KeyError: Key not in DB')


def write_to_xl_file(airport_data: dict) -> None:

    """ In Progress - currently DB is fine so only updates necessary """

    # Stamp the current time for document version control
    now = datetime.datetime.now()

    # Create a new Excel sheet to dump the data into
    file_path = os.getcwd() + "/MVM_Destinations_v" + now.strftime("%d%m%y%H%M") + ".xlsx"

    print('Writing info to', file_path)

    wb = openpyxl.Workbook()

    headers = ["iata", "name", "timezone", "latitude", "longitude", "DSTOffset", "DSTStart", "DSTEnd"]

    for key, value in airport_data.items():
        pass

    # Write new Excel workbook to disk..
    wb.save(file_path)


def update_dst(file_name) -> None:

    """
    Updates DST start and end dates in the inputted spreadsheet.

    :param file_name: file name with values that need updating
    :return: None
    """

    print('Updating spreadsheet DST information...')

    # Get the Excel file with the airport info, load the workbook and its sheet
    xlsx_file = os.getcwd() + '/' + file_name
    wb = load_workbook(filename=xlsx_file, data_only=True)
    master_sheet = wb['MASTER_LIST']

    json_db_data = open_json_db()

    # Collect info on each airport in the DB
    for row in master_sheet.iter_rows(row_offset=1):

        # Grab iata code from the master list
        iata = row[0].value

        if iata:

            dst_dates = json_db_data[iata][6]

            # Only write info if available
            if not dst_dates:
                continue
            elif len(dst_dates) < 2:
                # DST removed alltogether...
                row[6].value = dst_dates[0]
            else:
                row[6].value = dst_dates[0]
                row[7].value = dst_dates[1]

    wb.save(xlsx_file)


def main(args):

    if len(args) < 3:
        print('Not enough arguments, usage is: ./main.py [file_name_in_root_dir] [mode] [dst_year or iata_code]')
        exit()

    # File name, Input for the year you want to get DST start/end dates for, mode
    file_name, mode = args[1], args[2]

    if mode == '--dst':
        entry = input("Has the JSON DB been updated first? (y/n):")
        if entry.lower() != 'y':
            dst_year = args[3]
            data = get_airport_data(file_name, dst_year)
            update_json_db(data)
        update_dst(file_name)
    elif mode == '--new':
        dst_year = args[3]
        data = open_json_db()
        # write_to_xl_file(data)
    elif mode == '--get':
        iata = args[3].upper()
        get_from_db(iata)
    elif mode == '--update':
        dst_year = args[3]
        data = get_airport_data(file_name, dst_year)
        update_json_db(data)


if __name__ == '__main__':
    main(sys.argv)
